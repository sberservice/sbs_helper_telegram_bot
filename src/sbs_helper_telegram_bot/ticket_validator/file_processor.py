"""
Модуль обработки файлов для пакетной валидации заявок.

Обрабатывает чтение/запись Excel-файлов и пакетную валидацию заявок.
Поддерживает форматы .xlsx и устаревший .xls.
"""

import os
import re
import logging
from typing import List, Dict, Any, Optional, Callable, Tuple
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

# Regex для поиска недопустимых XML-символов (управляющие символы, кроме таба, перевода строки и возврата каретки)
ILLEGAL_XML_CHARS_RE = re.compile(
    '[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff\ufffe\uffff]'
)

# Символы, которые Excel воспринимает как начало формулы
FORMULA_START_CHARS = ('=', '+', '-', '@', '\t', '\r', '\n')


def sanitize_for_excel(value: Any) -> Any:
    """
    Подготовить значение для записи в Excel.
    
    Удаляет недопустимые XML-символы, которые портят Excel-файл.
    Экранирует строки, похожие на формулы, чтобы Excel не интерпретировал их как формулы.
    
    Args:
        value: Значение для очистки
        
    Returns:
        Очищенное значение, безопасное для Excel
    """
    if value is None:
        return value
    
    if isinstance(value, str):
        # Удаляем недопустимые XML-символы
        sanitized = ILLEGAL_XML_CHARS_RE.sub('', value)
        
        # Экранируем строки, похожие на формулы, добавляя ведущую одинарную кавычку
        # Excel покажет значение как текст (кавычка не отображается)
        if sanitized and sanitized[0] in FORMULA_START_CHARS:
            sanitized = "'" + sanitized
        
        # Также ограничиваем длину строки, чтобы избежать проблем (лимит ячейки Excel — 32767 символов)
        if len(sanitized) > 32000:
            sanitized = sanitized[:32000] + "... [ОБРЕЗАНО]"
        
        return sanitized
    
    return value


@dataclass
class TicketValidationRow:
    """Результат валидации одной заявки из файла"""
    row_number: int
    ticket_text: str
    is_valid: bool
    ticket_type: str
    errors: str
    passed_rules: List[str] = field(default_factory=list)
    original_row: List[Any] = field(default_factory=list)


@dataclass 
class FileValidationResult:
    """Результат пакетной валидации файла"""
    total_tickets: int
    valid_tickets: int
    invalid_tickets: int
    skipped_tickets: int
    results: List[TicketValidationRow]
    output_file_path: Optional[str] = None
    error_message: Optional[str] = None


class ExcelFileProcessor:
    """
    Обрабатывает Excel-файлы для пакетной валидации заявок.
    
    Поддерживает чтение .xlsx и .xls, валидацию всех заявок
    и запись результатов в новый Excel-файл.
    """
    
    def __init__(self):
        self.progress_callback: Optional[Callable[[int, int], None]] = None
        self._openpyxl = None
        self._xlrd = None
    
    def _ensure_openpyxl(self):
        """Ленивый импорт openpyxl для файлов xlsx"""
        if self._openpyxl is None:
            try:
                import openpyxl
                self._openpyxl = openpyxl
            except ImportError:
                raise ImportError(
                    "openpyxl is required for processing .xlsx files. "
                    "Install it with: pip install openpyxl"
                )
        return self._openpyxl
    
    def _ensure_xlrd(self):
        """Ленивый импорт xlrd для файлов xls"""
        if self._xlrd is None:
            try:
                import xlrd
                self._xlrd = xlrd
            except ImportError:
                raise ImportError(
                    "xlrd is required for processing .xls files. "
                    "Install it with: pip install xlrd"
                )
        return self._xlrd
    
    def read_file(self, file_path: str) -> Tuple[List[str], List[List[Any]]]:
        """
        Read Excel file and return headers and data.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (headers, rows)
            
        Raises:
            ValueError: Если формат файла не поддерживается
        """
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.xlsx':
            return self._read_xlsx(file_path)
        elif ext == '.xls':
            return self._read_xls(file_path)
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}. Используйте .xls или .xlsx")
    
    def _read_xlsx(self, file_path: str) -> Tuple[List[str], List[List[Any]]]:
        """Прочитать файл .xlsx с помощью openpyxl"""
        openpyxl = self._ensure_openpyxl()
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Получаем заголовки из первой строки
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
            
            # Получаем строки данных
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Пропускаем полностью пустые строки
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                rows.append(list(row))
            
            wb.close()
            return headers, rows
            
        except Exception as e:
            logger.error(f"Error reading xlsx file: {e}")
            raise ValueError(f"Error reading file: {str(e)}")
    
    def _read_xls(self, file_path: str) -> Tuple[List[str], List[List[Any]]]:
        """Прочитать устаревший файл .xls с помощью xlrd"""
        xlrd = self._ensure_xlrd()
        
        try:
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            
            # Получаем заголовки
            headers = []
            for col in range(ws.ncols):
                val = ws.cell_value(0, col)
                headers.append(str(val) if val else "")
            
            # Получаем строки данных
            rows = []
            for row_idx in range(1, ws.nrows):
                row = []
                all_empty = True
                for col in range(ws.ncols):
                    val = ws.cell_value(row_idx, col)
                    row.append(val)
                    if val and str(val).strip():
                        all_empty = False
                # Пропускаем полностью пустые строки
                if not all_empty:
                    rows.append(row)
            
            return headers, rows
            
        except Exception as e:
            logger.error(f"Error reading xls file: {e}")
            raise ValueError(f"Error reading file: {str(e)}")
    
    def get_column_names(self, file_path: str) -> List[str]:
        """
        Получить названия столбцов из Excel-файла.
        
        Args:
            file_path: Путь к Excel-файлу
            
        Returns:
            Список названий столбцов
        """
        headers, _ = self.read_file(file_path)
        return headers
    
    def validate_file(
        self,
        file_path: str,
        ticket_column: str | int,
        output_path: Optional[str] = None,
        ticket_type_id: Optional[int] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> FileValidationResult:
        """
        Провалидировать все заявки в файле.
        
        Args:
            file_path: Путь к входному Excel-файлу
            ticket_column: Название столбца или индекс (с 0), содержащий заявки
            output_path: Путь для выходного файла (None = сгенерировать автоматически)
            ticket_type_id: Принудительно задать тип заявки (None = автоопределение)
            progress_callback: Функция для уведомления о прогрессе (current, total)
        
        Returns:
            FileValidationResult с деталями валидации
        """
        # Импортируем модули валидации здесь, чтобы избежать циклических импортов
        from .validators import validate_ticket, detect_ticket_type, ValidationResult
        from .validation_rules import load_rules_from_db, load_all_ticket_types
        
        self.progress_callback = progress_callback
        
        try:
            # Читаем файл
            headers, rows = self.read_file(file_path)
            
            if not rows:
                return FileValidationResult(
                    total_tickets=0,
                    valid_tickets=0,
                    invalid_tickets=0,
                    skipped_tickets=0,
                    results=[],
                    error_message="Файл не содержит данных"
                )
            
            # Находим индекс столбца с заявками
            col_idx = self._resolve_column_index(headers, ticket_column)
            
            # Получаем типы заявок для автоопределения
            ticket_types = load_all_ticket_types() if ticket_type_id is None else None
            
            # Валидируем каждую заявку
            results = []
            valid_count = 0
            invalid_count = 0
            skipped_count = 0
            
            total_rows = len(rows)
            
            for idx, row in enumerate(rows, start=1):
                # Получаем текст заявки из строки
                if col_idx >= len(row):
                    ticket_text = ""
                else:
                    ticket_text = str(row[col_idx]) if row[col_idx] is not None else ""
                
                ticket_text = ticket_text.strip()
                
                # Пропускаем пустые заявки
                if not ticket_text:
                    result = TicketValidationRow(
                        row_number=idx,
                        ticket_text="",
                        is_valid=False,
                        ticket_type="Пустая строка",
                        errors="Текст заявки пуст",
                        passed_rules=[],
                        original_row=list(row)
                    )
                    results.append(result)
                    skipped_count += 1
                    
                    if self.progress_callback:
                        self.progress_callback(idx, total_rows)
                    continue
                
                # Определяем тип заявки, если он не задан явно
                detected_type = None
                type_id = ticket_type_id
                
                if ticket_type_id is None and ticket_types:
                    detected_type, _ = detect_ticket_type(ticket_text, ticket_types)
                    type_id = detected_type.id if detected_type else None
                elif ticket_type_id:
                    # Загружаем данные принудительно заданного типа заявки
                    from .validation_rules import load_ticket_type_by_id
                    detected_type = load_ticket_type_by_id(ticket_type_id)
                
                # Загружаем правила и валидируем
                if type_id:
                    rules = load_rules_from_db(type_id)
                    validation_result = validate_ticket(ticket_text, rules, detected_type)
                else:
                    # Не удалось определить тип
                    validation_result = ValidationResult(
                        is_valid=False,
                        failed_rules=[],
                        passed_rules=[],
                        error_messages=["Не удалось определить тип заявки"],
                        validation_details={},
                        detected_ticket_type=None
                    )
                
                # Сохраняем результат
                result = TicketValidationRow(
                    row_number=idx,
                    ticket_text=ticket_text[:500] + "..." if len(ticket_text) > 500 else ticket_text,
                    is_valid=validation_result.is_valid,
                    ticket_type=detected_type.type_name if detected_type else "Не определён",
                    errors="; ".join(validation_result.error_messages),
                    passed_rules=validation_result.passed_rules,
                    original_row=list(row)
                )
                results.append(result)
                
                if validation_result.is_valid:
                    valid_count += 1
                else:
                    invalid_count += 1
                
                # Коллбэк прогресса
                if self.progress_callback:
                    self.progress_callback(idx, total_rows)
            
            # Генерируем выходной файл
            if output_path is None:
                base_name = os.path.splitext(file_path)[0]
                output_path = f"{base_name}_validated.xlsx"
            
            self._write_results(headers, results, col_idx, output_path)
            
            return FileValidationResult(
                total_tickets=total_rows,
                valid_tickets=valid_count,
                invalid_tickets=invalid_count,
                skipped_tickets=skipped_count,
                results=results,
                output_file_path=output_path
            )
            
        except Exception as e:
            logger.error(f"Error validating file: {e}", exc_info=True)
            return FileValidationResult(
                total_tickets=0,
                valid_tickets=0,
                invalid_tickets=0,
                skipped_tickets=0,
                results=[],
                error_message=f"Ошибка обработки файла: {str(e)}"
            )
    
    def _resolve_column_index(self, headers: List[str], ticket_column: str | int) -> int:
        """
        Resolve column name or index to numeric index.
        
        Args:
            headers: Список заголовков столбцов
            ticket_column: Название столбца или индекс (с 0)
            
        Returns:
            Индекс столбца (с 0)
            
        Raises:
            ValueError: Если столбец не найден
        """
        if isinstance(ticket_column, int):
            if ticket_column < 0 or ticket_column >= len(headers):
                raise ValueError(f"Индекс столбца {ticket_column} вне диапазона (0-{len(headers)-1})")
            return ticket_column
        
        # Сначала пробуем точное совпадение
        if ticket_column in headers:
            return headers.index(ticket_column)
        
        # Пробуем совпадение без учёта регистра
        lower_headers = [h.lower() for h in headers]
        if ticket_column.lower() in lower_headers:
            return lower_headers.index(ticket_column.lower())
        
        raise ValueError(f"Столбец '{ticket_column}' не найден в файле. Доступны: {', '.join(headers)}")
    
    def _write_results(
        self,
        original_headers: List[str],
        results: List[TicketValidationRow],
        ticket_col_idx: int,
        output_path: str
    ):
        """
        Записать результаты валидации в Excel-файл.
        
        Args:
            original_headers: Исходные заголовки столбцов
            results: Список результатов валидации
            ticket_col_idx: Индекс столбца с заявкой
            output_path: Путь к выходному файлу
        """
        openpyxl = self._ensure_openpyxl()
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты валидации"
        
        # Определяем стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        valid_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        valid_font = Font(color="006100")
        invalid_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        invalid_font = Font(color="9C0006")
        skipped_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        skipped_font = Font(color="9C5700")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Записываем заголовки
        new_headers = original_headers + ["✅ Результат", "🎫 Тип заявки", "❌ Ошибки", "✓ Пройденные проверки"]
        for col_idx, header in enumerate(new_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Записываем данные
        for row_idx, result in enumerate(results, start=2):
            # Записываем исходные столбцы (очищаем все значения)
            for col_idx, value in enumerate(result.original_row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_for_excel(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Заполняем оставшиеся исходные столбцы, если строка короче
            for col_idx in range(len(result.original_row) + 1, len(original_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.border = thin_border
            
            # Столбец статуса валидации
            status_col = len(original_headers) + 1
            
            if result.ticket_type == "Пустая строка":
                status_text = "⏭️ ПРОПУЩЕН"
                fill = skipped_fill
                font = skipped_font
            elif result.is_valid:
                status_text = "✅ ВАЛИДНА"
                fill = valid_fill
                font = valid_font
            else:
                status_text = "❌ ОШИБКИ"
                fill = invalid_fill
                font = invalid_font
            
            status_cell = ws.cell(row=row_idx, column=status_col, value=status_text)
            status_cell.fill = fill
            status_cell.font = font
            status_cell.alignment = Alignment(horizontal='center', vertical='center')
            status_cell.border = thin_border
            
            # Столбец типа заявки
            type_cell = ws.cell(row=row_idx, column=status_col + 1, value=sanitize_for_excel(result.ticket_type))
            type_cell.border = thin_border
            type_cell.alignment = Alignment(vertical='top')
            
            # Столбец ошибок
            errors_cell = ws.cell(row=row_idx, column=status_col + 2, value=sanitize_for_excel(result.errors))
            errors_cell.border = thin_border
            errors_cell.alignment = Alignment(vertical='top', wrap_text=True)
            if result.errors:
                errors_cell.fill = invalid_fill
            
            # Столбец пройденных правил
            passed_text = "; ".join(result.passed_rules) if result.passed_rules else ""
            passed_cell = ws.cell(row=row_idx, column=status_col + 3, value=sanitize_for_excel(passed_text))
            passed_cell.border = thin_border
            passed_cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Автоподстройка ширины столбцов
        for col_idx in range(1, len(new_headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, len(results) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    # Учитываем переносы строк
                    if col_idx == len(original_headers) + 3:  # Столбец ошибок
                        cell_length = min(cell_length, 60)
                    elif col_idx == len(original_headers) + 4:  # Столбец пройденных правил
                        cell_length = min(cell_length, 50)
                    max_length = max(max_length, cell_length)
            
            # Ограничиваем ширину столбца
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Закрепляем строку заголовков
        ws.freeze_panes = 'A2'
        
        # Добавляем лист со сводкой
        self._add_summary_sheet(wb, results)
        
        wb.save(output_path)
        logger.info(f"Wrote validation results to {output_path}")
    
    def _add_summary_sheet(self, wb, results: List[TicketValidationRow]):
        """Добавить лист со статистикой"""
        ws = wb.create_sheet("Статистика")
        
        from openpyxl.styles import PatternFill, Font, Alignment
        
        header_font = Font(bold=True, size=14)
        
        total = len(results)
        valid = sum(1 for r in results if r.is_valid)
        invalid = sum(1 for r in results if not r.is_valid and r.ticket_type != "Пустая строка")
        skipped = sum(1 for r in results if r.ticket_type == "Пустая строка")
        
        # Записываем сводку
        ws.cell(row=1, column=1, value="Статистика валидации").font = header_font
        ws.cell(row=3, column=1, value="Всего строк:")
        ws.cell(row=3, column=2, value=total)
        ws.cell(row=4, column=1, value="✅ Валидных:")
        ws.cell(row=4, column=2, value=valid)
        ws.cell(row=5, column=1, value="❌ С ошибками:")
        ws.cell(row=5, column=2, value=invalid)
        ws.cell(row=6, column=1, value="⏭️ Пропущено:")
        ws.cell(row=6, column=2, value=skipped)
        
        if total > 0:
            success_rate = (valid / (total - skipped) * 100) if (total - skipped) > 0 else 0
            ws.cell(row=8, column=1, value="Процент успеха:")
            ws.cell(row=8, column=2, value=f"{success_rate:.1f}%")
        
        # Считаем ошибки по типам
        error_counts: Dict[str, int] = {}
        for result in results:
            if result.errors:
                for error in result.errors.split("; "):
                    error = error.strip()
                    if error:
                        error_counts[error] = error_counts.get(error, 0) + 1
        
        if error_counts:
            ws.cell(row=10, column=1, value="Распространённые ошибки:").font = header_font
            row = 12
            for error, count in sorted(error_counts.items(), key=lambda x: -x[1])[:10]:
                ws.cell(row=row, column=1, value=sanitize_for_excel(error))
                ws.cell(row=row, column=2, value=count)
                row += 1
        
        # Считаем по типам заявок
        type_counts: Dict[str, int] = {}
        for result in results:
            if result.ticket_type and result.ticket_type != "Пустая строка":
                type_counts[result.ticket_type] = type_counts.get(result.ticket_type, 0) + 1
        
        if type_counts:
            ws.cell(row=row + 2, column=1, value="Типы заявок:").font = header_font
            row += 4
            for ticket_type, count in sorted(type_counts.items(), key=lambda x: -x[1]):
                ws.cell(row=row, column=1, value=sanitize_for_excel(ticket_type))
                ws.cell(row=row, column=2, value=count)
                row += 1
        
        # Настраиваем ширину столбцов
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15


def get_column_names(file_path: str) -> List[str]:
    """
    Удобная функция для получения названий столбцов из Excel-файла.
    
    Args:
        file_path: Путь к Excel-файлу
        
    Returns:
        Список названий столбцов
    """
    processor = ExcelFileProcessor()
    return processor.get_column_names(file_path)


def validate_excel_file(
    file_path: str,
    ticket_column: str | int,
    output_path: Optional[str] = None,
    ticket_type_id: Optional[int] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None
) -> FileValidationResult:
    """
    Convenience function to validate tickets in an Excel file.
    
    Args:
        file_path: Path to input Excel file
        ticket_column: Column name or index containing tickets
        output_path: Path for output file (None = auto-generate)
        ticket_type_id: Force specific ticket type (None = auto-detect)
        progress_callback: Function to call with progress (current, total)
    
    Returns:
        FileValidationResult with validation details
    """
    processor = ExcelFileProcessor()
    return processor.validate_file(
        file_path=file_path,
        ticket_column=ticket_column,
        output_path=output_path,
        ticket_type_id=ticket_type_id,
        progress_callback=progress_callback
    )
